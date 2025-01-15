import time
from openpyxl import load_workbook
from config import base_path, fonte1_filename, fonte2_filename
from config import base_tabname, fonte1_tabname, fonte2_tabname


def merge():
    print("loading Base...")

    base_workbook = load_workbook(filename=base_path)
    base = base_workbook[base_tabname]

    print("loading Fonte 1...")
    fonte1_workbook = load_workbook(filename=fonte1_filename) 
    fonte1 = fonte1_workbook[fonte1_tabname]

    print("loading Fonte 2...")
    fonte2_workbook = load_workbook(filename="info/FONTE 2.xlsx")
    fonte2 = fonte2_workbook["Sheet1"]

    print("workbooks loaded successfully.")
    
    total_rows = len(fonte1['C'])
    print(f"Iniciando processamento de {total_rows - 1} linhas da Fonte 1")

    for fonte1_row, fonte1_cell in enumerate(fonte1['C']):
        if fonte1_row == 0:
            continue

        code = fonte1_cell.value
        if code is None:
            print(f"FONTE 1: Linha {fonte1_row + 1} sem Código. Pulando...")
            continue
    
        target_row = 0
        for base_row, base_cell in enumerate(base['D']):
            if base_cell.value and base_cell.value == code:
                target_row = base_row + 1
                break

        if target_row == 0:
            print(f"FONTE 1: Código '{code}' não encontrado na BASE. Adicionando nova linha.")
            target_row = base.max_row + 1
        else:
            print(f"FONTE 1: Código '{code}' encotrado. Atualizando a linha {target_row}.")

        new_number = fonte1.cell(row=fonte1_row+1, column=1).value
        new_date = fonte1.cell(row=fonte1_row+1, column=2).value
        new_value1 = fonte1.cell(row=fonte1_row+1, column=4).value

        base[f"A{target_row}"] = new_number
        base[f"B{target_row}"] = "Polo SC"
        base[f"C{target_row}"] = new_date
        base[f"D{target_row}"] = code
        base[f"E{target_row}"] = new_value1

    base_workbook.save("info/BASE.xlsx")

    total_rows = len(fonte2['C'])
    print(f"Iniciando processamento de {total_rows - 1} linhas da Fonte 2")

    for fonte2_row, fonte2_cell in enumerate(fonte2['C']):
        if fonte2_row == 0:
            continue

        code = fonte2_cell.value
        if code is None:
            print(f"FONTE 2: Linha {fonte2_row + 1} sem Código. Pulando...")
            continue
        
        target_row = 0
        for base_row, base_cell in enumerate(base['D']):
            if base_cell.value and base_cell.value == code:
                target_row = base_row + 1
                break
        if target_row == 0:
            print(f"FONTE 2: Código '{code}' não encontrado na BASE. Adicionando nova linha.")
            target_row = base.max_row + 1
        else:
            print(f"FONTE 2: Código '{code}' encotrado. Atualizando a linha {target_row}.")

        new_number = fonte2.cell(row=fonte2_row+1, column=1).value
        new_date2 = fonte2.cell(row=fonte2_row+1, column=2).value

        new_text = f"IMPORTAÇÃO NF {new_number} DE {new_date2.strftime("%d/%m/%Y")}"
        new_value2 = fonte2.cell(row=fonte2_row+1, column=5).value
        new_ipi = fonte2.cell(row=fonte2_row+1, column=6).value

        base[f"F{target_row}"] = new_value2
        base[f"G{target_row}"] = new_ipi
        base[f"H{target_row}"] = new_text

        print("A Base foi atualizada com Sucesso.")

    return 0
